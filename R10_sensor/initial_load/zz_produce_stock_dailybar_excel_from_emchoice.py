import pandas as pd
from pandas import Series, DataFrame
import numpy as np
import urllib.error
from bs4 import BeautifulSoup
import re
from datetime import datetime

import time
import R50_general.advanced_helper_funcs as ahf
import R50_general.general_constants
import R50_general.general_helper_funcs as gcf
from R50_general.general_helper_funcs import logprint,get_tmp_file
import R50_general.dfm_to_table_common as df2db
import openpyxl

global_module_name = gcf.get_cur_file_name_by_module_name(__name__)

max_stockids_per_excel = 10
excel_output_path = R50_general.general_constants.emchoice_excel_output_path
process_time_per_excel = 120000
em_server_local_path = 'C:\working\\'   # the local path of em choice server which script to run

ls_pars = [
            'PRECLOSE',	#前收盘价
            'OPEN',	#开盘价
            'HIGH',	#最高价
            'LOW',	#最低价
            'CLOSE',	#收盘价
            'CHANGE',	#涨跌
            'PCTCHANGE',	#涨跌幅
            'VOLUME',	#成交量
            'AMOUNT',	#成交额
            'AVERAGE',	#均价
            'TURN',	#换手率
            'AMPLITUDE',	#振幅
            'TRADESTATUS',	#交易状态
            'TNUM',	#成交笔数
            'TAFACTOR',	#复权因子(后)
            'BUYVOL',	#内盘成交量
            'SELLVOL',	#外盘成交量
            'HIGHLIMIT',	#是否涨停
            'LOWLIMIT',	#是否跌停
            'ISSTSTOCK',	#是否ST
            'ISXSTSTOCK',	#是否*ST
]

ls_filenames =[]

def produce_excels(stockid:str = ''):

    # step1.1: get current stock list
    dfm_stocks = df2db.get_cn_stocklist(stockid,include_inactive=False)
    # print(dfm_stocks)

    start_date = get_start_date()
    end_date = get_end_date()

    str_pars = 'ref!A1:A%s' %len(ls_pars)

    stock_no = 0
    ls_stocks =[]
    file_id = 0
    for index,row in dfm_stocks.iterrows():
        ls_stocks.append(row['Stock_ID']+'.'+row['Market_ID'])
        stock_no +=1
        if stock_no == max_stockids_per_excel:
            stock_no = 0
            file_id +=1
            make_excel(ls_stocks,str_pars,start_date,end_date,file_id)
            ls_stocks =[]

    if stock_no > 0:
        file_id+=1
        make_excel(ls_stocks, str_pars, start_date, end_date,file_id)

    make_batch_process_vbs()

def make_excel(ls_stocks, str_pars, start_date, end_date,file_id):

    str_stocks = 'ref!B1:B%s' %len(ls_stocks)

    str_formula = build_em_hq(str_stocks, str_pars, start_date, end_date)
    # print(len(str_formula))
    # print(str_formula)
    file_name = 'fetch_dailybars_%s_to_%s_%s.xlsx' %(start_date,end_date,file_id)
    logprint('Produce excel %s for stocks %s' % (file_name,ls_stocks))
    wb = openpyxl.Workbook()
    sheet_hist = wb.create_sheet(index=0, title='hist')
    sheet_ref = wb.create_sheet(index=1, title='ref')
    sheet_hist['A1'] = str_formula
    for i in range(len(ls_pars)):
        sheet_ref['A' + str(i+1)] = ls_pars[i]
    for i in range(len(ls_stocks)):
        sheet_ref['B' + str(i+1)] = ls_stocks[i]

    filename_with_path  = excel_output_path + file_name
    wb.save(filename_with_path)

    ls_filenames.append(file_name)


def build_em_hq(stocks,pars,start_date,end_date):
    '''
 =EM_HQ(历史行情参数页1!B1:B500,    stock list
        历史行情参数页2!A1:A21,     pars to be fetch
        "2005-01-01",             start date
        "2018-03-29",             end date
        "Period=1,                dailybar
        AdjustFlag=1,             no adjust
        Type=1,
        Layout1=-1,Layout2=-1,
        Order=0,                  sort ascd
        DateType=0,               date = 'YYYY-MM-DD'
        Market=CNSESH,            market
        ClearArea=NULL,
        basedate=N")
    :return:
    '''
    str_formula = '''=EM_HQ(%s,%s,"%s","%s","Period=1,AdjustFlag=1,Type=1,Layout1=-1,Layout2=-1,Order=0,DateType=0,Market=CNSESH,ClearArea=NULL,basedate=N")''' %(
        stocks,pars,start_date,end_date)
    return str_formula


def get_start_date():
    return '2005-01-01'

def get_end_date():
    return '2018-03-29'

def make_batch_process_vbs():

    logprint('Creating batch process vbs script...')
    batch_vbs = open(excel_output_path+'batch_open_save_excels.vbs','w')
    str_vbs_open_excel_app = '''Set objExcel = CreateObject("Excel.Application") 
objExcel.Visible = True
objExcel.AddIns2.Item("EMFunc").Installed = False
objExcel.AddIns2.Item("EMFunc").Installed = True
    '''

    str_vbs_process_excel = '''Set objWorkbook = objExcel.Workbooks.Open ("%s") 
set WshShell = WScript.CreateObject("WScript.Shell")   
WScript.Sleep %s
objExcel.Workbooks(1).Save '保存工作表 
objExcel.Workbooks(1).Close '关闭工作表 
    '''

    str_vbs_exit_excel_app = '''objExcel.Quit  ' 退出 
'wscript.echo "处理完成"
    '''

    batch_vbs.write(str_vbs_open_excel_app)

    for filename in ls_filenames:
        batch_vbs.write(str_vbs_process_excel %(em_server_local_path+filename,process_time_per_excel))

    batch_vbs.write(str_vbs_exit_excel_app)

    batch_vbs.close()

if __name__ == '__main__':
    produce_excels('')


# sample script
r'''
Set objExcel = CreateObject("Excel.Application") 
objExcel.Visible = True
    
objExcel.AddIns2.Item("EMFunc").Installed = False
objExcel.AddIns2.Item("EMFunc").Installed = True

Set objWorkbook = objExcel.Workbooks.Open ("C:\working\fetch_dailybars_2005-01-01_to_2018-03-29_1.xlsx") 


set WshShell = WScript.CreateObject("WScript.Shell")   
WScript.Sleep 60000


objExcel.Workbooks(1).Save '保存工作表 
objExcel.Workbooks(1).Close '关闭工作表 

Set objWorkbook = objExcel.Workbooks.Open ("C:\working\fetch_dailybars_2005-01-01_to_2018-03-29_2.xlsx") 

set WshShell = WScript.CreateObject("WScript.Shell")   
WScript.Sleep 60000


objExcel.Workbooks(1).Save '保存工作表 
objExcel.Workbooks(1).Close '关闭工作表 

objExcel.Quit  ' 退出 
'wscript.echo "保存成功"
'''